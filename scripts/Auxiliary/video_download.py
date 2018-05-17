import urllib.parse
import urllib.request
from bs4 import BeautifulSoup
import pytube
import os
import youtube_dl
from openpyxl import load_workbook
from xlrd import open_workbook
from xlutils.copy import copy


wb=open_workbook('./Songs_Database_300.xlsx')
# nb=copy(wb)#copy the original workbook in a new workboook
# ns=nb.get_sheet(0)
s=wb.sheet_by_index(0);
for i in range(300,301):
	textToSearch = 'Trippy Trippy (Bhoomi)Laagi Na Choote (A Gentleman)'
	#textToSearch = s.cell(i,1).value
	query = urllib.parse.quote(textToSearch,safe='')
	url = "https://www.youtube.com/results?search_query=" + query
	response = urllib.request.urlopen(url)
	html = response.read()
	soup = BeautifulSoup(html,"lxml")
	song_url=''
	for vid in soup.findAll(attrs={'class':'yt-uix-tile-link'}):
	    song_url=('https://www.youtube.com' + vid['href'])
	    break
	print(song_url)


	song_url = s.cell(i,3).value
	print(song_url)
	if __name__=='__main__':
		try:
			yt=pytube.YouTube(song_url);
			#print('1');
			# print(yt.filename())
			# yt.set_filename(textToSearch)

			#print(yt.streams.first())
			path='./dataset_videos';
			if(not os.path.exists(path)):
				os.makedirs(path)
				#print('inside');
			#print('1');
			yt.streams.first().download(path,filename=textToSearch);
			print('done  '+textToSearch)
			#print('1');
			# videos=yt.get_videos()
			# for v in videos:
			# 	print(str(v))

			# print("done")	
			# #video=yt.get('mp4','720p')
			# path="D:\AMAN\Capstone"

			# video[0].download(path)
		except Exception as e:
			print(e);



# from youtube_search_and_download import YouTubeHandler
 
# search_key = 'chinese top ktv' #keywords
# yy = YouTubeHandler(search_key)
# yy.download_as_audio =0 # 1- download as audio format, 0 - download as video
# yy.set_num_playlist_to_extract(3) # number of playlist to download
 
# print ('Get all the playlist')
# yy.get_playlist_url_list()
# print (yy.playlist_url_list)
 
# ## Get all the individual video and title from each of the playlist
# yy.get_video_link_fr_all_playlist()
# for key in  yy.video_link_title_dict.keys():
#     print (key, '  ', yy.video_link_title_dict[key])
    
# print ('download video')
# yy.download_all_videos(dl_limit =200) #number of videos to download.













# 	artist=s.cell(17+i,1).value
# 	song=s.cell(17+i,2).value
# 	try:
# 		a=PyLyrics.getLyrics(artist,song)
# 	except:
# 		a=''
# 	ns.write(17+i,4,a)
# print('done')#(name of artist,name of song)
# nb.save('newsheet1.xls')